import openpyxl
import traceback
import os
import json
import qrcode

# Define variable to load the dataframe

def exception_to_string(exception):
    try:
        text = str(exception)
        text += '\n'
        stack = traceback.format_tb(tb=exception.__traceback__)
        stack.reverse()
        for entry in stack:
            terms = entry.split(',')
            head, filename = os.path.split(terms[0])
            text += filename.replace('"', '') + terms[1].replace('line ', ':').replace(' ', '')
            text += terms[2].replace('\n', '')
            text += '\n'

        return text

    except Exception as e:
        return 'Failed to process exception!'


def do_stuff():

    data = {}

    #workbook = openpyxl.load_workbook('game.xlsx')
    workbook = openpyxl.load_workbook('game_my_interventions.xlsx')

    for sheet in workbook.worksheets:
        print(sheet)

        data[sheet.title] = []

        index = 0

        for row in range(2, sheet.max_row+1):

            record = {}
            for column in range(1, sheet.max_column+1):
                label = sheet.cell(row=1, column=column).value
                value = sheet.cell(row=row, column=column).value

                if value == None:
                    value = ""

                if isinstance(value, int):
                    value = str(value)

                if isinstance(value, float):
                    value = str(value)

                value = value.replace('\n', ' <br> ')
                value = value.replace('<b>', ' <b> ')
                value = value.replace('</b>', ' </b> ')
                value = value.replace('\u00a0', ' ')
                value = value.replace('\u2019',"'")
                value = value.replace('\u2018', "'")

                record[label] = value

            if sheet.title != 'text':
                record['url'] = 'https://arsinoe-serious-game.github.io/proto-2/website/game/'+ sheet.title +'/' + sheet.title+'_'+str(index)+'.html'
                record['local_url'] = 'website/game/' + sheet.title + '/' + sheet.title + '_' + str(index) + '.html'

                img = qrcode.make(record['url'])
                type(img)  # qrcode.image.pil.PilImage

                record['qr_code'] = 'assets/'+sheet.title+'/' + sheet.title+'_'+str(index)+'_qr.png'
                img.save('../' + record['qr_code'])

                if sheet.title == 'interventions':
                    with open('../website/game/' + sheet.title + '/' + sheet.title + '_' + str(index) + '.html','w') as fp:
                        """
                        <!DOCTYPE html>
                        <html data-bs-theme="light" lang="en">
                        
                        <head>
                            <meta charset="utf-8">
                            <meta name="viewport" content="width=device-width, initial-scale=1.0, shrink-to-fit=no">
                            <title>BP:Approve a fish farm</title>
                            <link rel="stylesheet" href="../../../assets/bootstrap/css/bootstrap.min.css">
                        </head>
                        
                        <body>
                        <header>
                            <h1 align="center">Approve a fish farm:interventions_0.html </h1>
                        </header>
                            <div id="page_content" style="width: 90%;margin: 5%;">
                            <p><span style="color: rgb(0, 0, 0);">Sed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem aperiam, eaque ipsa quae ab illo inventore veritatis et quasi architecto beatae vitae dicta sunt explicabo. Nemo enim ipsam voluptatem quia voluptas sit aspernatur aut odit aut fugit, sed quia consequuntur magni dolores eos qui ratione voluptatem sequi nesciunt. Neque porro quisquam est, qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit, sed quia non numquam eius modi tempora incidunt ut labore et dolore magnam aliquam quaerat voluptatem. Ut enim ad minima veniam, quis nostrum exercitationem ullam corporis suscipit laboriosam, nisi ut aliquid ex ea commodi consequatur? Quis autem vel eum iure reprehenderit qui in ea voluptate velit esse quam nihil molestiae consequatur, vel illum qui dolorem eum fugiat quo voluptas nulla pariatur?</span></p>
                            </div>
                        
                            <div style="text-align: center">
                                <img src="../../../assets/cards/interventions/interventions-0.png" width="75%" >
                            </div>
                        
                            <div id="page_content-2" style="width: 90%;margin: 5%;">
                            <p><span style="color: rgb(0, 0, 0);">Sed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem aperiam, eaque ipsa quae ab illo inventore veritatis et quasi architecto beatae vitae dicta sunt explicabo. Nemo enim ipsam voluptatem quia voluptas sit aspernatur aut odit aut fugit, sed quia consequuntur magni dolores eos qui ratione voluptatem sequi nesciunt. Neque porro quisquam est, qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit, sed quia non numquam eius modi tempora incidunt ut labore et dolore magnam aliquam quaerat voluptatem. Ut enim ad minima veniam, quis nostrum exercitationem ullam corporis suscipit laboriosam, nisi ut aliquid ex ea commodi consequatur? Quis autem vel eum iure reprehenderit qui in ea voluptate velit esse quam nihil molestiae consequatur, vel illum qui dolorem eum fugiat quo voluptas nulla pariatur?</span></p>
                            </div>
                        
                            <div style="text-align: center">
                                <div class="d-grid gap-2 ms-1">
                                    <a class="btn btn-primary" role="button" href="../../../interventions.html" style="width: 33%;margin: 0px;margin-left: 33%;" display="inline-block">Home</a>
                                </div>
                                <p><br></p>
                                <p><br></p>
                                <p><br></p>
                                <p><br></p>
                            </div>
                            <script src="../../../assets/bootstrap/js/bootstrap.min.js"></script>
                        </body>
                        </html>                        
                        """

                        fp.write('<!DOCTYPE html>')
                        fp.write('<html data-bs-theme="light" lang="en">')

                        fp.write('<head>')
                        fp.write('<meta charset="utf-8">')
                        fp.write('<meta name="viewport" content="width=device-width, initial-scale=1.0, shrink-to-fit=no">')
                        fp.write('<title>' + record['type'] + ':' + record['name'] + '</title>')
                        fp.write('<link rel="stylesheet" href="../../../assets/bootstrap/css/bootstrap.min.css">')
                        fp.write('</head>')

                        fp.write('<body>')
                        fp.write('<header>')
                        fp.write('<h1 align="center">')
                        fp.write(record['name'] + ' ' + sheet.title + '_' + str(index) + '.html')
                        fp.write('</h1>')
                        fp.write('</header>')
                        fp.write('<div id="page_content" style="width: 90%;margin: 5%;">')
                        fp.write('<p><span style="color: rgb(0, 0, 0);">Sed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem aperiam, eaque ipsa quae ab illo inventore veritatis et quasi architecto beatae vitae dicta sunt explicabo. Nemo enim ipsam voluptatem quia voluptas sit aspernatur aut odit aut fugit, sed quia consequuntur magni dolores eos qui ratione voluptatem sequi nesciunt. Neque porro quisquam est, qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit, sed quia non numquam eius modi tempora incidunt ut labore et dolore magnam aliquam quaerat voluptatem. Ut enim ad minima veniam, quis nostrum exercitationem ullam corporis suscipit laboriosam, nisi ut aliquid ex ea commodi consequatur? Quis autem vel eum iure reprehenderit qui in ea voluptate velit esse quam nihil molestiae consequatur, vel illum qui dolorem eum fugiat quo voluptas nulla pariatur?</span></p>')
                        fp.write('</div>')

                        fp.write('<div style="text-align: center">')
                        fp.write('<img src="../../../assets/cards/' + sheet.title + '/' + sheet.title+'-'+str(index)+'.png" width="75%" >')
                        fp.write('</div>')

                        fp.write('<div id="page_content-2" style="width: 90%;margin: 5%;">')
                        fp.write('<p><span style="color: rgb(0, 0, 0);">Sed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem aperiam, eaque ipsa quae ab illo inventore veritatis et quasi architecto beatae vitae dicta sunt explicabo. Nemo enim ipsam voluptatem quia voluptas sit aspernatur aut odit aut fugit, sed quia consequuntur magni dolores eos qui ratione voluptatem sequi nesciunt. Neque porro quisquam est, qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit, sed quia non numquam eius modi tempora incidunt ut labore et dolore magnam aliquam quaerat voluptatem. Ut enim ad minima veniam, quis nostrum exercitationem ullam corporis suscipit laboriosam, nisi ut aliquid ex ea commodi consequatur? Quis autem vel eum iure reprehenderit qui in ea voluptate velit esse quam nihil molestiae consequatur, vel illum qui dolorem eum fugiat quo voluptas nulla pariatur?</span></p>')
                        fp.write('</div>')

                        fp.write('<div style="text-align: center">')
                        fp.write('<div class="d-grid gap-2 ms-1">')
                        fp.write('<a class="btn btn-primary" role="button" href="../../../interventions.html" style="width: 33%;margin: 0px;margin-left: 33%;" display="inline-block">Home</a>')
                        fp.write('</div>')
                        fp.write('<p><br></p>')
                        fp.write('<p><br></p>')
                        fp.write('<p><br></p>')
                        fp.write('<p><br></p>')
                        fp.write('</div>')
                        fp.write('<script src="../../../assets/bootstrap/js/bootstrap.min.js"></script>')
                        fp.write('</body>')
                        fp.write('</html>')
                else:
                    #make some html content for page
                    with open('../website/game/'+sheet.title+'/' + sheet.title+'_'+str(index)+'.html', 'w') as fp:
                        fp.write('<H1>\n')
                        fp.write(record['name'])
                        fp.write('\n')
                        fp.write('\n')
                        fp.write(sheet.title + '_' + str(index) + '.html')
                        fp.write('\n')
                        fp.write('</H1>\n')
                        fp.write('<img src="../../../assets/cards/' + sheet.title +'/' + sheet.title+'-'+str(index)+'.png">')
                        fp.close()

                index += 1

            data[sheet.title].append(record)

    with open('../game/game_data.js', 'w') as fp:
        fp.write('let game_data = ')
        fp.write('\n')
        json.dump(data, fp, indent=4)
        fp.write(';\n')
        fp.write('\n')
        fp.close()

    return


do_stuff()
