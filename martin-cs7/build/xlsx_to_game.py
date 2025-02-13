import openpyxl
import traceback
import os
import json
import qrcode

# Define variable to load the dataframe

target_url = 'https://arsinoe-serious-game.github.io/martin-cs7/'


def exception_to_string(exception):
    try:
        text = str(exception)
        text += '\n'
        stack = traceback.format_tb(tb=exception.__traceback__)
        stack.reverse()
        for entry in stack:
            terms = entry.split(',')
            head, filename = os.path.split(terms[0])
            text += filename.replace('"', '') + terms[1].replace('line ', ':').replace(' ', '')
            text += terms[2].replace('\n', '')
            text += '\n'

        return text

    except Exception as e:
        return 'Failed to process exception!'


def do_stuff():

    data = {}

    #workbook = openpyxl.load_workbook('game.xlsx')
    workbook = openpyxl.load_workbook('game_my_interventions.xlsx')

    for sheet in workbook.worksheets:
        print(sheet)

        data[sheet.title] = []

        index = 0

        for row in range(2, sheet.max_row+1):

            record = {}
            for column in range(1, sheet.max_column+1):
                label = sheet.cell(row=1, column=column).value
                value = sheet.cell(row=row, column=column).value

                if value == None:
                    value = ""

                if isinstance(value, int):
                    value = str(value)

                if isinstance(value, float):
                    value = str(value)

                value = value.replace('\n', ' <br> ')
                value = value.replace('<b>', ' <b> ')
                value = value.replace('</b>', ' </b> ')
                value = value.replace('\u00a0', ' ')
                value = value.replace('\u2019',"'")
                value = value.replace('\u2018', "'")

                record[label] = value

            if sheet.title != 'text':
                number_label = str(index).zfill(2)

                if index == 0:
                    number_label = '0'

                record['local_url'] = 'kb/' + sheet.title + '/' + sheet.title + '_' + number_label + '.html'

                if sheet.title == 'persona':
                    record['local_url'] = 'kb/' + sheet.title + '/' + 'personas.html'

                record['url'] = target_url + record['local_url']


                img = qrcode.make(record['url'])
                type(img)  # qrcode.image.pil.PilImage

                if not os.path.exists('game/'+ 'assets/'+sheet.title+'/'):
                    os.makedirs('game/'+ 'assets/'+sheet.title+'/')

                record['qr_code'] = 'assets/'+sheet.title+'/' + sheet.title+'_' + number_label + '_qr.png'
                img.save('game/'+ record['qr_code'])

                #just have 1 persona
                if sheet.title != 'persona':
                    index += 1

            data[sheet.title].append(record)

    with open('game/game_data.js', 'w') as fp:
        fp.write('let game_data = ')
        fp.write('\n')
        json.dump(data, fp, indent=4)
        fp.write(';\n')
        fp.write('\n')
        fp.close()

    return


do_stuff()
